"""
Tools collection for comparing article separations
"""

import logging, sys
from pathlib import Path, PurePath
import pickle
from sqlite3 import connect
from csv import DictWriter, DictReader
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side

from python_util.parser.xml.page.page import Page

logger = logging.getLogger(__name__)


class SeparatedPage(Page):
    """container class for PAGE-XML with article annotation per baseline"""

    def __init__(self, xmlFilePath: Path):
        super(SeparatedPage, self).__init__(str(xmlFilePath))
        self.xmlFilePath = xmlFilePath
        self._blIgnore = set()
        self._reInit()

    def _reInit(self):
        self.blNiDict = dict()
        self.niBlDict = {ni: [] for ni in self.get_article_dict().keys()}
        for bl in self.get_textlines():
            if not bl.id in self._blIgnore:
                niId = bl.get_article_id()
                blId = bl.id
                self.blNiDict[blId] = niId
                self.niBlDict[niId].append(blId)
        for (ni,niBl) in self.niBlDict.items():
            self.niBlDict[ni] = sorted(niBl)
        self._canBlPart = None

    def removeBlSet(self, blSet: set) -> None:
        """remove redundant baselines from partition"""
        self._blIgnore.update(blSet)
        self._reInit()

    def canonicalBlPartition(self) -> list:
        """canonical article-partition of baselines"""
        if self._canBlPart is None:
            self._canBlPart = sorted([sorted([bl.id for bl in blList]) for blList in self.get_article_dict().values()])
        return self._canBlPart


class SepPageComparison():
    """container class for comparison counters"""

    def __init__(self):
        self.gtNIs = None
        self.hypNIs = None
        self.corrects = None
        self.splits = None
        self.merges = None
        self.dist = None

    def __str__(self):
        return str(self.__dict__)

    def dataDict(self) -> dict:
        """returns data as dict"""
        return self.__dict__

    def loadDict(self, dataDict: dict) -> None:
        """loads data from dict"""
        for member in self.__dict__:
            setattr(self, member, int(dataDict.get(member, None)))

    def checkConsistency(self):
        return self.gtNIs + self.splits + self.merges == self.hypNIs


class SepPageComper():
    """comparision engines root class"""

    def __init__(self):
        self._hypSepPage = None
        self._gtSepPage = None
        self._altGtDict = {}
        self.comparison = None

    def loadGT(self, xmlFilePath: Path) -> None:
        """loads ground truth"""
        self._gtSepPage = SeparatedPage(xmlFilePath=xmlFilePath)

    def compareTo(self, xmlFilePath: Path) -> SepPageComparison:
        """runs comparison to hypothesis"""
        self._hypSepPage = SeparatedPage(xmlFilePath=xmlFilePath)
        self.comparison = self._compare()
        return self.comparison

    def _compare(self) -> SepPageComparison:
        """internal comparison procedure – to be implemented"""
        raise NotImplementedError('Method not yet been implemented!')


class SepPageBlComper(SepPageComper):
    """comparison engine based on baseline partitions"""

    def _compare(self) -> SepPageComparison:
        hypSepPage: SeparatedPage = self._hypSepPage
        hypBlSet = set(textLine.id for textLine in hypSepPage.get_textlines())
        actGtSepPage: SeparatedPage = self._gtSepPage
        gtBlSet = set(textLine.id for textLine in actGtSepPage.get_textlines())
        if gtBlSet != hypBlSet:
            if gtBlSet.issubset(hypBlSet):
                raise AssertionError(f'cannot compare: inconsistent baselines')
            diffSet = gtBlSet.difference(hypBlSet)
            gtBlSet = gtBlSet.difference(diffSet)
            altLabel = ''.join(diffSet)
            logger.debug(f'ignoring inconsistent baselines {altLabel}')
            actGtSepPage = self._altGtDict.get(altLabel, None)
            if actGtSepPage is None:
                actGtSepPage = SeparatedPage(self._gtSepPage.xmlFilePath)
                actGtSepPage.removeBlSet(diffSet)
                self._altGtDict[altLabel] = actGtSepPage
                logger.log(logging.DEBUG // 2, f'new SeparatedPage constructed for {altLabel}')
            else:
                logger.log(logging.DEBUG // 2, f'used SeparatedPage from dict')
            logger.log(logging.DEBUG // 2, f'altGtDict has size {len(self._altGtDict)}')

        comparison = SepPageComparison()
        comparison.gtNIs = len(actGtSepPage.niBlDict)
        comparison.hypNIs = len(hypSepPage.niBlDict)
        correct = [block for block in actGtSepPage.canonicalBlPartition() if
                   block in hypSepPage.canonicalBlPartition()]
        comparison.corrects = len(correct)
        inf = []
        infMmbr = set()
        for blID in gtBlSet:
            if not (blID in infMmbr):
                gtSet = set(actGtSepPage.niBlDict[actGtSepPage.blNiDict[blID]])
                hypSet = set(hypSepPage.niBlDict[hypSepPage.blNiDict[blID]])
                infSet = gtSet.intersection(hypSet)
                inf.append(list(infSet))
                infMmbr.update(infSet)
        comparison.splits = len(inf) - comparison.gtNIs
        comparison.merges = comparison.hypNIs - len(inf)
        comparison.dist = comparison.splits - comparison.merges
        return comparison


class SepPageCompDict(dict):
    """container class for comparison results"""
    fieldNames = ['dataSet', 'method', 'gtXML', 'hypXML', *SepPageComparison().dataDict().keys()]

    @classmethod
    def path2method(cls, path: str) -> str:
        """method extraction: last-but-one part of (file)path"""
        return str(PurePath(path).parent.parts[-1])

    def addItem(self, dataSet: str, gtXML: str, hypXML: str, spcDict: SepPageComparison) -> None:
        """add entry to collection"""
        self[dataSet] = self.get(dataSet, {})
        self[dataSet][gtXML] = self[dataSet].get(gtXML, {})
        self[dataSet][gtXML][hypXML] = spcDict

    def loadPickle(self, dataSetLabel: str, pickleFilePath: Path) -> None:
        """loading collection from pickle file"""
        with pickleFilePath.open(mode='rb') as pickleFile:
            self[dataSetLabel] = pickle.load(file=pickleFile)

    def cleanup(self, inclList: list) -> None:
        """removing entries of methods not in inclusion list"""
        for dataDict in self.values():
            for gtDict in dataDict.values():
                for hypString in gtDict:
                    if not SepPageCompDict.path2method(hypString) in inclList:
                        gtDict[hypString] = None

    def loadCSV(self, csvFilePath: Path, inclList: list) -> None:
        """loading collection from CSV file"""
        with csvFilePath.open(mode='rt') as csvFile:
            csvReader = DictReader(csvFile)
            for dataDict in csvReader:
                if dataDict.get('method').lower() in inclList:
                    if not dataDict.get('dataSet') in self:
                        dataSet = dataDict.get('dataSet')
                        self[dataSet] = {}
                    setDict = self[dataSet]
                    if not dataDict.get('gtXML') in setDict:
                        gtXML = dataDict.get('gtXML')
                        setDict[gtXML] = {}
                    gtDict = setDict[gtXML]
                    spc = SepPageComparison()
                    spc.loadDict(dataDict)
                    gtDict[dataDict.get('hypXML')] = spc

    def expSqlite(self, dbFilePath: Path, dbTableName: str):
        """exporting collection to SQLITE database"""
        fieldString = ', '.join(SepPageCompDict.fieldNames)
        dbCon = connect(dbFilePath)
        dbCur = dbCon.cursor()
        try:
            sqlCmd = f'DROP TABLE {dbTableName}'
            dbCur.execute(sqlCmd)
        except:
            pass
        sqlCmd = f'CREATE TABLE {dbTableName} ({fieldString})'
        dbCur.execute(sqlCmd)

        for (dataSet, dataDict) in self.items():
            for (gtXML, gtDict) in dataDict.items():
                for (hypXML, comp) in gtDict.items():
                    method = SepPageCompDict.path2method(hypXML)
                    dataList = [f'"{dataSet}"', f'"{method}"', f'"{gtXML}"', f'"{hypXML}"']
                    for (field, data) in comp.dataDict().items():
                        dataList.append(str(data))
                    dataString = ', '.join(dataList)
                    sqlCmd = f"INSERT INTO allComps ({fieldString}) VALUES ({dataString})"
                    dbCur.execute(sqlCmd)

        dbCon.commit()
        dbCon.close()

    def expCsv(self, csvFilePath: Path):
        """exporting collection to CSV file"""
        with csvFilePath.open(mode='wt', encoding='utf8', newline='') as csvFile:
            csvWriter = DictWriter(csvFile, fieldnames=SepPageCompDict.fieldNames)
            csvWriter.writeheader()
            for (dataSet, dataDict) in self.items():
                for (gtXML, gtDict) in dataDict.items():
                    for (hypXML, comp) in gtDict.items():
                        method = SepPageCompDict.path2method(hypXML)
                        dataDict = {'dataSet': dataSet, 'method': method, 'gtXML': gtXML,
                                    'hypXML': hypXML}
                        dataDict.update(comp.dataDict())
                        csvWriter.writerow(dataDict)


class CompDictEvaler():
    """engine for evaluating separation comparison result collections

    see README.md for further explanation
    """

    def __init__(self, spcDict: SepPageCompDict):
        self.spcDict = spcDict
        self.winnerStatDict = {}
        self.winnerDict = {}

    def countWinnerStat(self) -> None:
        """computes the statistics of victories"""
        for (dataSet, dataDict) in self.spcDict.items():
            self.winnerStatDict[dataSet] = {}
            statDict = self.winnerStatDict[dataSet]
            for (gtXML, gtDict) in dataDict.items():
                for (hypXML0, compDict0) in gtDict.items():
                    if compDict0:
                        method0 = SepPageCompDict.path2method(hypXML0)
                        if not method0 in statDict.keys():
                            statDict[method0] = {'all': 0}
                        for (hypXML1, compDict1) in gtDict.items():
                            if compDict1:
                                method1 = SepPageCompDict.path2method(hypXML1)
                                if not method1 in statDict[method0].keys():
                                    statDict[method0][method1] = 0
                                if (compDict0.dist, -compDict0.corrects) <= (compDict1.dist, -compDict1.corrects):
                                    statDict[method0][method1] += 1
                                    statDict[method0]['all'] += 1

    def calcWinnerDict(self) -> None:
        """collects winner statistics"""
        if len(self.winnerStatDict) == 0:
            self.countWinnerStat()
        for (dataSet, dataDict) in self.winnerStatDict.items():
            self.winnerDict[dataSet] = {}
            actDict = self.winnerDict[dataSet]
            actMethods = dataDict.keys()
            for method in actMethods:
                actDict[method] = [dataDict[method]['all']]
            actMethods = list(sorted(actMethods, key=lambda method: actDict[method][-1]))
            actDict['_max'] = [actDict[actMethods[-1]][-1]]
            while len(actMethods) > 1:
                actLooser = actMethods.pop(0)
                for method in actMethods:
                    actDict[method].append(actDict[method][-1] - self.winnerStatDict[dataSet][method][actLooser])
                actMethods = list(sorted(actMethods, key=lambda method: actDict[method][-1]))
                actDict['_max'].append(actDict[actMethods[-1]][-1])

    def winnerStat2xlsx(self, xlsxFilePath: Path) -> None:
        """exports winner statistics into XLSX workbook"""
        wb = Workbook()
        if len(self.winnerStatDict) > 0:
            for (dataSet, dataDict) in self.winnerStatDict.items():
                methodList = list(sorted(dataDict.keys(), key=lambda method: dataDict[method]['all'], reverse=True))
                wb.create_sheet(dataSet)
                ws = wb.get_sheet_by_name(dataSet)
                actCell = ws.cell(row=1, column=1)
                actCell.value = "all"
                actCell.font = Font(bold=True)
                actCell.alignment = Alignment(horizontal='center')
                col = 2
                for method1 in methodList:
                    col += 1
                    actCell = ws.cell(row=1, column=col)
                    actCell.value = method1
                    actCell.font = Font(bold=True)
                    actCell.alignment = Alignment(horizontal='center')
                maxCol = col
                for col in range(1, maxCol + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 40
                row = 1
                for method0 in methodList:
                    row += 1
                    actCell = ws.cell(row=row, column=1)
                    actCell.value = dataDict[method0]['all']
                    actCell.alignment = Alignment(horizontal='center')
                    actCell = ws.cell(row=row, column=2)
                    actCell.value = method0
                    actCell.font = Font(bold=True)
                    actCell.alignment = Alignment(horizontal='center')
                    for col in range(3, maxCol + 1):
                        method1 = ws.cell(row=1, column=col).value
                        actCell = ws.cell(row=row, column=col)
                        actCell.alignment = Alignment(horizontal='center')
                        if method0 == method1:
                            actCell.value = dataDict[method0][method1]
                            actCell.font = Font(color='666666')
                        else:
                            denom = dataDict[method1][method0]
                            if denom > 0:
                                ratio = dataDict[method0][method1] / denom
                                actCell.value = ratio
                                actCell.number_format = '0.00'
                                if ratio < 1.0:
                                    actCell.font = Font(color='880000')
                                else:
                                    actCell.font = Font(color='00DD00')
                            else:
                                actCell.font = Font(color='00DD00')

        if len(self.winnerDict) > 0:
            wb.create_sheet('winner', index=0)
            ws = wb.get_sheet_by_name('winner')
            ws.column_dimensions['A'].width = 40
            rowOffset = 0
            for (dataSet, dataDict) in self.winnerDict.items():
                methodList = [method for method in dataDict.keys() if not method.startswith('_')]
                methodList = list(sorted(methodList, key=lambda method: len(dataDict[method]), reverse=True))
                row = 1
                actCell = ws.cell(row=rowOffset + row, column=1)
                actCell.value = dataSet
                actCell.alignment = Alignment(horizontal='left')
                actSide = Side(border_style='medium')
                actCell.border = Border(left=actSide, right=actSide, top=actSide, bottom=actSide)
                for method in methodList:
                    row += 1
                    actCell = ws.cell(row=rowOffset + row, column=1)
                    actCell.value = method
                    actCell.font = Font(bold=True)
                    actCell.alignment = Alignment(horizontal='center')
                    colOffset = 2
                    for (index, value) in enumerate(dataDict[method]):
                        actCell = ws.cell(row=rowOffset + row, column=colOffset + index)
                        actCell.value = value
                        if value == dataDict['_max'][index]:
                            actCell.font = Font(bold=True)
                        actCell.alignment = Alignment(horizontal='center')
                rowOffset += row + 1

        try:
            wb.remove(wb.get_sheet_by_name('Sheet'))
        except:
            pass
        wb.save(xlsxFilePath)
